import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PIL import Image
import io
import tempfile
import datetime

st.set_page_config(page_title="Generador Manual - Guía de Medición", layout="centered")
st.title("📏 Generador Manual de Guía de Medición")
st.write("Subí captura, completá los campos y descargá el Excel con la guía (formato: Screenshot | how it is triggered | Script | Variable | Values).")

with st.form("manual_form"):
    st.subheader("1) Subí la captura del evento (botón, zona, etc.)")
    uploaded_image = st.file_uploader("Imagen (.png, .jpg)", type=["png", "jpg", "jpeg"])

    st.subheader("2) Cómo se dispara")
    how_triggered = st.text_input("Describe cómo se dispara el evento (ej: On Click → Button Comprar)")

    st.subheader("3) Parámetros del evento (escribilos vos)")
    # event selection for uaevent / nievent
    interaction = st.radio("¿Tiene interacción?", ("Sí (uaevent)", "No (nievent)"))
    event_val = "uaevent" if interaction.startswith("Sí") else "nievent"

    eventCategory = st.text_input("eventCategory", value="")
    eventAction = st.text_input("eventAction", value="")
    eventLabel = st.text_input("eventLabel", value="")
    event_name = st.text_input("event_name", value="")

    st.markdown("---")
    submit = st.form_submit_button("Generar Excel")

def build_script_block(event, category, action, label, name):
    # arma un bloque tipo dataLayer.push con los campos
    lines = [
        "dataLayer.push({",
        f'  event: "{event}",',
        f'  eventCategory: "{category}",',
        f'  eventAction: "{action}",',
        f'  eventLabel: "{label}",',
        f'  event_name: "{name}"',
        "});"
    ]
    return "\n".join(lines)

def create_excel_bytes(img_file, how_triggered_text, script_text,
                       variables_dict, filename_prefix="guia"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Guía de Medición"

    # Cabeceras
    headers = ["Screenshot", "how it is triggered", "Script", "Variable", "Values"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    # Vamos a reservar 5 filas para las variables (event + 4 restantes)
    var_rows = max(5, len(variables_dict))
    start_row = 2
    end_row = start_row + var_rows - 1

    # Merge para Screenshot | how it is triggered | Script para que se alineen con las filas de variables
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # A2:A(n)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)  # B2:B(n)
    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)  # C2:C(n)

    # Insertar how it is triggered
    ws.cell(row=start_row, column=2, value=how_triggered_text)

    # Insertar script (en la celda C2)
    ws.cell(row=start_row, column=3, value=script_text)

    # Insertar variables y values en columnas D y E desde row=2 hacia abajo
    r = start_row
    for var, val in variables_dict.items():
        ws.cell(row=r, column=4, value=var)
        ws.cell(row=r, column=5, value=val)
        r += 1

    # Ajustes visuales: ancho columnas
    widths = {1: 25, 2: 35, 3: 50, 4: 20, 5: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Si hay imagen, insertarla en A2 (ya merged)
    if img_file is not None:
        # aseguramos tener un archivo temporario para openpyxl Image
        try:
            # Si img_file es un BytesIO o file-like, lo guardamos temporalmente
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                # Abrir con PIL y redimensionar para que entre bien (max ancho 600px)
                image = Image.open(img_file)
                max_width = 800
                ratio = min(1.0, max_width / image.width)
                new_size = (int(image.width * ratio), int(image.height * ratio))
                image = image.convert("RGBA")
                image.thumbnail(new_size, Image.ANTIALIAS)
                image.save(tmp.name, format="PNG")

                img = XLImage(tmp.name)
                # colocar la imagen anclada en A2
                img.anchor = "A2"
                ws.add_image(img)
        except Exception as e:
            print("Error insertando imagen en Excel:", e)

    # Guardar workbook a bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"{filename_prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return bio, filename

if submit:
    # validar uploads mínimos
    if not uploaded_image:
        st.warning("Subí al menos una imagen para incluir en la guía (puede ser un screenshot del botón).")
    else:
        st.success("Generando Excel...")
        # armar script
        script = build_script_block(event_val, eventCategory, eventAction, eventLabel, event_name)

        # variables ordenadas (para que siempre salga event primero)
        variables = {
            "event": event_val,
            "eventCategory": eventCategory,
            "eventAction": eventAction,
            "eventLabel": eventLabel,
            "event_name": event_name
        }

        # crear excel en bytes
        # uploaded_image es un UploadedFile; para PIL lo convertimos a BytesIO
        img_bytes = io.BytesIO(uploaded_image.getvalue())
        excel_bytes_io, out_name = create_excel_bytes(img_bytes, how_triggered, script, variables, filename_prefix="guia_medicion")

        st.download_button("⬇ Descargar Excel (.xlsx)", excel_bytes_io, file_name=out_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.balloons()

